import urllib.request as res
import re
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import requests



#编写正则
def getjobname(jobname,startnum,endnum):
    allresult=[]
    jobname1=res.quote(res.quote(jobname))
    for i in range(startnum,endnum):
        print('正在抓取第',i,'页')
        url='https://search.51job.com/list/000000,000000,0000,00,9,99,'+jobname1+',2,'+str(i)+'.html?lang=c&stype=1&postchannel=0000&workyear=99&cotype=99&degreefrom=99&jobterm=99&companysize=99&lonlat=0%2C0&radius=-1&ord_field=0&confirmdate=9&fromType=4&dibiaoid=0&address=&line=&specialarea=00&from=&welfare='
        response=res.urlopen(url)
        html=response.read().decode('gbk')
        #编写正则用来提取信息，去掉的东西都用.*?来代替
        reg=re.compile('<div class="el">.*?<p class="t1 ">.*?<a target="_blank" title="(.*?)".*?<span class="t2"><a target="_blank" title="(.*?)".*?<span class="t3">(.*?)</span>.*?<span class="t4">(.*?)</span>.*?<span class="t5">(.*?)</span>.*?</div>',re.S)
        result=re.findall(reg,html)
        allresult += result


#存储到excel
    if os.path.exists('51job职位信息.xlsx'):
        workbook = load_workbook('51job职位信息.xlsx')
        sheetNames = workbook.sheetnames
        if jobname in sheetNames:
            print('sheetname 重复')
            return
        else:
            sheet = workbook.create_sheet(jobname,len(sheetNames)+1)
            sheet.append('职位名,公司名,工作地点,薪资,发布时间'.split(','))
            for each in allresult:
                sheet.append(each)
    else:
        workbook =Workbook()
        sheet =workbook.active
        sheet.title =jobname
        sheet.append('职位名,公司名,工作地点,薪资,发布时间'.split(','))
        for each in allresult:
            sheet.append(each)
    workbook.save('51job职位信息.xlsx')





getjobname('数据分析', 1, 100) #执行函数